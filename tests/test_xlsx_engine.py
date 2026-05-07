import csv
import json
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from minial_agent.common.utils import file_registry
from minial_agent.integrations.upload import ensure_upload_workspace
from minial_agent.integrations.upload.registry import UploadRegistry
from minial_agent.integrations.xlsx.dataframes import profile_dataframe
from minial_agent.integrations.xlsx.detect_tables import detect_table
from minial_agent.integrations.xlsx.errors import XlsxRangeError, XlsxTransformError
from minial_agent.integrations.xlsx.exports import commit_workbook, export_detected_table_csv, export_range
from minial_agent.integrations.xlsx.ranges import range_to_dataframe, read_range
from minial_agent.integrations.xlsx.sessions import XlsxSessionStore
from minial_agent.integrations.xlsx.transforms import transform_dataframe
from minial_agent.integrations.xlsx.workbook import inspect_workbook


@dataclass(frozen=True)
class TableWorkbookSpec:
    sheet_name: str = "Data"
    header_row: int = 3
    start_col: int = 2
    data_rows: int = 8
    headers: tuple[str, ...] = (
        "Field 1",
        "Field 2",
        "Group",
        "Score",
        "Field 5",
        "Field 6",
    )
    banner_rows: tuple[int, ...] = ()

    @property
    def end_col(self) -> int:
        return self.start_col + len(self.headers) - 1

    @property
    def end_row(self) -> int:
        return self.header_row + self.data_rows

    @property
    def table_range(self) -> str:
        return (
            f"{get_column_letter(self.start_col)}{self.header_row}:"
            f"{get_column_letter(self.end_col)}{self.end_row}"
        )


def test_xlsx_engine_inspects_workbook_and_reads_range(tmp_path) -> None:
    spec = TableWorkbookSpec()
    path = tmp_path / "table.xlsx"
    write_plain_table_workbook(path, spec)

    inspection = inspect_workbook(path).to_dict()
    assert inspection["sheet_count"] == 3
    assert inspection["sheets"][0]["sheet_name"] == spec.sheet_name
    assert inspection["sheets"][0]["used_range"] == spec.table_range
    assert inspection["sheets"][0]["has_tables"] is True
    assert inspection["sheets"][2]["visible"] is False

    result = read_range(path, spec.sheet_name, spec.table_range)
    assert result.headers == list(spec.headers)
    assert result.to_dict()["column_count"] == len(spec.headers)
    assert result.to_dict()["row_count"] == spec.data_rows
    assert result.rows[0][spec.headers[0]] == generated_row_values(0)[0]


def test_xlsx_engine_rejects_invalid_sheet_and_range(tmp_path) -> None:
    spec = TableWorkbookSpec()
    path = tmp_path / "table.xlsx"
    write_plain_table_workbook(path, spec)

    with pytest.raises(XlsxRangeError, match="sheet not found"):
        read_range(path, "Missing", spec.table_range)

    with pytest.raises(ValueError):
        read_range(path, spec.sheet_name, "bad")


def test_xlsx_engine_exports_xlsx_and_csv(tmp_path, monkeypatch) -> None:
    spec = TableWorkbookSpec()
    workspace = ensure_upload_workspace(tmp_path / "workspace")
    source_path = workspace.files_dir / "table.xlsx"
    write_plain_table_workbook(source_path, spec)
    register_xlsx(workspace, source_path)
    stub_artifact_builder(monkeypatch)
    artifact = file_registry.resolve_artifact(
        workspace=workspace,
        file_ref="file_001",
        expected_file_type="xlsx",
    )

    xlsx_result = export_range(
        workspace=workspace,
        source_artifact=artifact,
        workbook_path=source_path,
        sheet=spec.sheet_name,
        range_ref=spec.table_range,
        output_path="/table_export.xlsx",
    )
    assert xlsx_result["file"]["file_id"] == "file_002"
    exported = load_workbook(workspace.files_dir / "table_export.xlsx")
    try:
        assert [
            exported.active.cell(1, column).value
            for column in range(1, len(spec.headers) + 1)
        ] == list(spec.headers)
        assert exported.active.max_row == spec.data_rows + 1
    finally:
        exported.close()

    csv_result = export_range(
        workspace=workspace,
        source_artifact=artifact,
        workbook_path=source_path,
        sheet=spec.sheet_name,
        range_ref=spec.table_range,
        output_path="/table_export.csv",
    )
    assert csv_result["file"]["filename"] == "table_export.csv"
    rows = read_csv(workspace.files_dir / "table_export.csv")
    assert rows[0] == list(spec.headers)
    assert len(rows) == spec.data_rows + 1

    with pytest.raises(XlsxRangeError, match="Output path"):
        export_range(
            workspace=workspace,
            source_artifact=artifact,
            workbook_path=source_path,
            sheet=spec.sheet_name,
            range_ref=spec.table_range,
            output_path="/table_export.txt",
        )


def test_xlsx_engine_detects_bannered_table_and_exports_csv(tmp_path) -> None:
    spec = TableWorkbookSpec(header_row=6, data_rows=10, banner_rows=(2, 4))
    workspace = ensure_upload_workspace(tmp_path / "workspace")
    source_path = workspace.files_dir / "bannered_table.xlsx"
    write_bannered_table_workbook(source_path, spec)
    (workspace.files_dir / "extracted.csv").write_text("bad,data\n", encoding="utf-8")

    table = detect_table(source_path)
    assert table.sheet == spec.sheet_name
    assert table.range == spec.table_range
    assert table.headers == list(spec.headers)
    assert table.row_count == spec.data_rows
    assert table.column_count == len(spec.headers)

    result = export_detected_table_csv(
        workspace=workspace,
        workbook_path=source_path,
        output_filename="extracted.csv",
    )
    assert result["file"]["filename"] == "extracted.csv"
    assert result["file"]["visible_path"] == "/extracted.csv"
    assert result["overwritten"] is True
    assert result["detected_range"]["range"] == spec.table_range
    assert result["detected_range"]["row_count"] == spec.data_rows
    rows = read_csv(workspace.files_dir / "extracted.csv")
    assert rows[0] == list(spec.headers)
    assert rows[1] == [str(value) for value in generated_row_values(0)]
    assert len(rows) == spec.data_rows + 1
    assert banner_values(spec).isdisjoint({row[0] for row in rows})
    assert "nan" not in "\n".join(",".join(row) for row in rows)


def test_xlsx_engine_csv_result_handles_relative_workspace_path(tmp_path, monkeypatch) -> None:
    spec = TableWorkbookSpec(header_row=6, banner_rows=(2, 4))
    monkeypatch.chdir(tmp_path)
    workspace = ensure_upload_workspace(Path("workspace"))
    source_path = workspace.files_dir / "bannered_table.xlsx"
    write_bannered_table_workbook(source_path, spec)

    result = export_detected_table_csv(
        workspace=workspace,
        workbook_path=source_path,
        output_filename="extracted.csv",
    )

    assert result["file"]["visible_path"] == "/extracted.csv"
    assert (workspace.files_dir / "extracted.csv").is_file()


def test_xlsx_engine_csv_filename_forms_and_internal_path_rejection(tmp_path) -> None:
    spec = TableWorkbookSpec(header_row=6, banner_rows=(2, 4))
    workspace = ensure_upload_workspace(tmp_path / "workspace")
    source_path = workspace.files_dir / "table.xlsx"
    write_bannered_table_workbook(source_path, spec)

    for requested in (
        "plain_name",
        "/rooted.csv",
        "files/prefixed.csv",
        "/workspace/files/legacy.csv",
    ):
        result = export_detected_table_csv(
            workspace=workspace,
            workbook_path=source_path,
            output_filename=requested,
        )
        assert (workspace.files_dir / result["file"]["visible_path"].lstrip("/")).is_file()
        assert result["file"]["filename"].endswith(".csv")

    for requested in (".jobs/out.csv", ".converted/out.csv", "../out.csv"):
        with pytest.raises(ValueError):
            export_detected_table_csv(
                workspace=workspace,
                workbook_path=source_path,
                output_filename=requested,
            )


def test_xlsx_engine_profiles_and_restricts_transforms(tmp_path) -> None:
    spec = TableWorkbookSpec()
    path = tmp_path / "table.xlsx"
    write_plain_table_workbook(path, spec)
    dataframe = range_to_dataframe(path, spec.sheet_name, spec.table_range)
    profile = profile_dataframe("table", dataframe).to_dict()

    assert profile["shape"] == [spec.data_rows, len(spec.headers)]
    assert profile["columns"] == list(spec.headers)
    assert profile["numeric_stats"]["Score"]["count"] == spec.data_rows

    input_path = tmp_path / "input.json"
    output_path = tmp_path / "output.json"
    from minial_agent.integrations.xlsx.dataframes import save_dataframe

    save_dataframe(input_path, dataframe)
    transformed = transform_dataframe(
        input_path=input_path,
        output_path=output_path,
        code=(
            "def transform(df):\n"
            "    return df.groupby('Group', as_index=False)['Score'].mean().sort_values('Score')\n"
        ),
    )
    assert list(transformed.columns) == ["Group", "Score"]

    with pytest.raises(XlsxTransformError, match="Imports are not allowed"):
        transform_dataframe(
            input_path=input_path,
            output_path=output_path,
            code="import os\ndef transform(df):\n    return df\n",
        )


def test_xlsx_session_runs_edit_formula_export_and_commit(tmp_path, monkeypatch) -> None:
    spec = TableWorkbookSpec()
    workspace = ensure_upload_workspace(tmp_path / "workspace")
    source_path = workspace.files_dir / "table.xlsx"
    write_plain_table_workbook(source_path, spec)
    register_xlsx(workspace, source_path)
    stub_artifact_builder(monkeypatch)
    artifact = file_registry.resolve_artifact(
        workspace=workspace,
        file_ref="file_001",
        expected_file_type="xlsx",
    )

    session = XlsxSessionStore(workspace).create(artifact=artifact, instruction="summarize table")
    loaded = session.load_range(
        sheet=spec.sheet_name,
        range_ref=spec.table_range,
        dataframe_name="table",
        header=True,
    )
    assert loaded["profile"]["shape"] == [spec.data_rows, len(spec.headers)]

    transformed = session.transform(
        input_dataframe="table",
        output_dataframe="summary",
        explanation="Average score by group",
        code=(
            "def transform(df):\n"
            "    return df.groupby('Group', as_index=False)['Score'].mean()\n"
        ),
    )
    assert transformed["dataframe"] == "summary"
    session.write_dataframe(
        dataframe_name="summary",
        sheet="Summary",
        start_cell="A1",
        include_header=True,
    )
    formula_result = session.add_formula(
        sheet="Summary",
        cell="D1",
        formula="=AVERAGE(B2:B4)",
        fill_range=None,
    )
    assert formula_result["changed_count"] == 1

    commit = commit_workbook(
        workspace=workspace,
        source_artifact=artifact,
        workbook_path=session.working_path,
        output_path="/table_summary.xlsx",
        summary="Added table summary.",
        changed_items=session.changed_items(),
    )
    assert commit["file"]["file_id"] == "file_002"
    assert source_path.is_file()
    assert (workspace.jobs_dir / "xlsx_sessions" / session.manifest.session_id / "manifest.json").is_file()
    assert ".jobs" not in json.dumps(commit)


def test_xlsx_imports_do_not_cycle() -> None:
    subprocess.run(
        [
            sys.executable,
            "-c",
            "from minial_agent.integrations.xlsx.workbook import inspect_workbook",
        ],
        check=True,
    )
    subprocess.run(
        [
            sys.executable,
            "-c",
            (
                "from minial_agent.integrations.upload.xlsx import build_xlsx_artifacts; "
                "from minial_agent.integrations.xlsx.workbook import inspect_workbook"
            ),
        ],
        check=True,
    )


def write_plain_table_workbook(path: Path, spec: TableWorkbookSpec) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.sheet_name
    write_table(sheet, spec)
    table = Table(displayName="DataTable", ref=spec.table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Summary"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()


def write_bannered_table_workbook(path: Path, spec: TableWorkbookSpec) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.sheet_name
    for row in spec.banner_rows:
        sheet.cell(row=row, column=spec.start_col).value = f"Banner Row {row}"
    write_table(sheet, spec)
    workbook.save(path)
    workbook.close()


def write_table(sheet, spec: TableWorkbookSpec) -> None:
    for col_index, header in enumerate(spec.headers, start=spec.start_col):
        sheet.cell(row=spec.header_row, column=col_index).value = header
    for offset in range(spec.data_rows):
        row = spec.header_row + 1 + offset
        for col_index, value in enumerate(generated_row_values(offset), start=spec.start_col):
            sheet.cell(row=row, column=col_index).value = value


def generated_row_values(offset: int) -> list[Any]:
    return [
        f"item-{offset + 1}",
        f"label-{offset + 1}",
        "A" if offset % 2 == 0 else "B",
        offset + 1,
        f"text-{offset + 1}",
        100 + offset,
    ]


def banner_values(spec: TableWorkbookSpec) -> set[str]:
    return {f"Banner Row {row}" for row in spec.banner_rows}


def read_csv(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8", newline="") as file:
        return list(csv.reader(file))


def register_xlsx(workspace, source_path: Path) -> None:
    converted_dir = workspace.converted_dir / "file_001"
    converted_dir.mkdir(parents=True)
    (converted_dir / "manifest.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": source_path.name,
                "source_path": str(source_path),
                "file_type": "xlsx",
                "pages": [],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name=source_path.name,
        file_type="xlsx",
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")


def stub_artifact_builder(monkeypatch) -> None:
    def fake_build_upload_artifacts(**kwargs):
        target_dir = kwargs["converted_dir"]
        target_dir.mkdir(parents=True, exist_ok=True)
        (target_dir / "manifest.json").write_text(
            json.dumps(
                {
                    "file_id": kwargs["file_id"],
                    "source_filename": kwargs["source_path"].name,
                    "file_type": kwargs["file_type"],
                    "pages": [],
                    "status": "converted",
                }
            ),
            encoding="utf-8",
        )

    monkeypatch.setattr(file_registry, "build_upload_artifacts", fake_build_upload_artifacts)
