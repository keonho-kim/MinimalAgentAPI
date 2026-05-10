import base64
import importlib
import json

from deepagents.backends.filesystem import FilesystemBackend
from deepagents.middleware.skills import _list_skills
from openpyxl import Workbook, load_workbook
import pytest

from minial_agent.agents.tools.read_documents.docx import (
    build_docx_read_workflow,
)
from minial_agent.agents.domain.office_file_editor.subagents.editor_xlsx.agent import (
    build_xlsx_subagent,
)
from minial_agent.agents.tools.read_documents.xlsx import (
    build_xlsx_read_workflow,
)
from minial_agent.agents.utils import (
    scan as scan_module,
)
from minial_agent.common.utils import file_registry
from minial_agent.integrations.upload import ensure_upload_workspace
from minial_agent.integrations.upload.registry import UploadRegistry
from minial_agent.integrations.upload.resolver import resolve_upload_artifact
from minial_agent.integrations.upload.visibility import (
    WorkspaceVisibilityError,
    normalize_public_workspace_path,
    to_public_workspace_path,
)
from minial_agent.integrations.xlsx.exports import commit_workbook
from minial_agent.integrations.xlsx.sessions import XlsxSessionStore

read_documents_module = importlib.import_module("minial_agent.agents.tools.read_documents")


def test_normalize_public_workspace_path_rejects_internal_paths() -> None:
    assert normalize_public_workspace_path("/workspace/files/report.docx") == (
        "/report.docx"
    )
    assert normalize_public_workspace_path("report.docx") == "/report.docx"
    assert to_public_workspace_path("/result.docx") == "files/result.docx"

    for hidden_path in (
        "/workspace/.registry/files.json",
        ".agents/skills/toy-skill/SKILL.md",
        ".converted/file_001/manifest.json",
        "/files/.secret",
        "/workspace/outputs/result.docx",
    ):
        try:
            normalize_public_workspace_path(hidden_path)
        except WorkspaceVisibilityError:
            continue
        raise AssertionError(f"Expected hidden path rejection: {hidden_path}")


def test_filesystem_backend_rooted_at_files_hides_internal_directories(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    (workspace.files_dir / "report.txt").write_text("hello", encoding="utf-8")
    (workspace.registry_dir / "secret.txt").write_text("secret", encoding="utf-8")

    backend = FilesystemBackend(
        root_dir=workspace.files_dir,
        virtual_mode=True,
        max_file_size_mb=1024,
    )

    root_entries = backend.ls("/").entries or []
    assert [entry["path"] for entry in root_entries] == ["/report.txt"]
    assert backend.read("/report.txt").error is None
    assert backend.read("/.registry/secret.txt").error is not None
    assert backend.grep("secret", "/").matches == []
    glob_matches = backend.glob("**/*", "/").matches or []
    assert [match["path"] for match in glob_matches] == ["/report.txt"]


def test_filesystem_backend_agent_paths_are_rooted_at_files(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    backend = FilesystemBackend(
        root_dir=workspace.files_dir,
        virtual_mode=True,
        max_file_size_mb=1024,
    )

    assert backend.write("/summary.md", "ok").error is None

    assert (workspace.files_dir / "summary.md").read_text(encoding="utf-8") == "ok"
    assert not (workspace.files_dir / "files" / "summary.md").exists()


def test_workspace_agents_skills_are_internal_and_loadable(tmp_path) -> None:
    skill_dir = tmp_path / ".agents" / "skills" / "toy-skill"
    skill_dir.mkdir(parents=True)
    (skill_dir / "SKILL.md").write_text(
        """---
name: toy-skill
description: Use this toy skill when testing workspace skill recognition.
---

# Toy Skill
""",
        encoding="utf-8",
    )

    workspace = ensure_upload_workspace(tmp_path)

    assert (workspace.skills_dir / "toy-skill" / "SKILL.md").is_file()
    assert not (workspace.files_dir / ".agents").exists()

    backend = FilesystemBackend(
        root_dir=workspace.agents_dir,
        virtual_mode=True,
        max_file_size_mb=1024,
    )
    skills = _list_skills(backend, "/skills")

    assert [skill["name"] for skill in skills] == ["toy-skill"]
    assert skills[0]["path"] == "/skills/toy-skill/SKILL.md"


def test_resolver_returns_public_metadata_without_internal_paths(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    source_path = workspace.files_dir / "report.docx"
    source_path.write_text("doc", encoding="utf-8")
    converted_dir = workspace.converted_dir / "file_001"
    converted_dir.mkdir()
    manifest_path = converted_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "report.docx",
                "source_path": str(source_path),
                "file_type": "docx",
                "pdf_path": str(converted_dir / "source.pdf"),
                "pages": [
                    {
                        "page_number": 1,
                        "image_filename": "page_001.png",
                        "image_path": str(converted_dir / "pages" / "page_001.png"),
                    }
                ],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name="report.docx",
        file_type="docx",
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")

    artifact = resolve_upload_artifact(
        workspace=workspace,
        file_ref="/workspace/files/report.docx",
        expected_file_type="docx",
    )
    metadata = artifact.public_metadata()

    assert metadata == {
        "file_id": "file_001",
        "filename": "report.docx",
        "file_type": "docx",
        "status": "converted",
        "visible_path": "files/report.docx",
        "page_count": 1,
    }
    assert ".converted" not in json.dumps(metadata)
    assert ".registry" not in json.dumps(metadata)


def test_workflow_result_does_not_expose_internal_paths(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(
        workspace,
        filename="report.docx",
        file_type="docx",
        page_count=3,
    )
    scanned_pages = []

    workflow = build_docx_read_workflow(
        workspace,
        page_scanner=lambda path, _question: (
            scanned_pages.append(path.name) or "matched page"
            if path.name == "page_002.png"
            else scanned_pages.append(path.name) or "None"
        ),
        evidence_judge=lambda _question, _evidence: True,
    )
    result = workflow.invoke(
        {
            "file_ref": "file_001",
            "question": "summary?",
        }
    )
    assert result["result"] == "page_2: matched page"
    assert result["evidence"] == {"page_2": "matched page"}
    assert result["scanned_pages"] == 3
    assert result["is_sufficient"] is True
    assert set(scanned_pages) == {"page_001.png", "page_002.png", "page_003.png"}
    assert ".converted" not in result["result"]
    assert ".registry" not in result["result"]


def test_workflow_stops_after_sufficient_page_scan_batch(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("AGENT_PAGE_SCAN_BATCH_SIZE", "10")
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(
        workspace,
        filename="report.docx",
        file_type="docx",
        page_count=25,
    )
    scanned_pages = []

    def page_scanner(path, _question):
        scanned_pages.append(path.name)
        return "matched page" if path.name == "page_012.png" else "None"

    workflow = build_docx_read_workflow(
        workspace,
        page_scanner=page_scanner,
        evidence_judge=lambda _question, _evidence: True,
    )

    result = workflow.invoke({"file_ref": "file_001", "question": "summary?"})

    assert result["result"] == "page_12: matched page"
    assert result["scanned_pages"] == 20
    assert "page_021.png" not in scanned_pages


def test_workflow_uses_configured_page_scan_batch_size(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("AGENT_PAGE_SCAN_BATCH_SIZE", "5")
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(
        workspace,
        filename="report.docx",
        file_type="docx",
        page_count=12,
    )
    scanned_pages = []

    def page_scanner(path, _question):
        scanned_pages.append(path.name)
        return "matched page" if path.name == "page_007.png" else "None"

    workflow = build_docx_read_workflow(
        workspace,
        page_scanner=page_scanner,
        evidence_judge=lambda _question, _evidence: True,
    )

    result = workflow.invoke({"file_ref": "file_001", "question": "summary?"})

    assert result["result"] == "page_7: matched page"
    assert result["scanned_pages"] == 10
    assert "page_011.png" not in scanned_pages


def test_workflow_full_scan_reads_all_pages(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("AGENT_PAGE_SCAN_BATCH_SIZE", "10")
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(
        workspace,
        filename="report.docx",
        file_type="docx",
        page_count=25,
    )
    scanned_pages = []

    def page_scanner(path, _question):
        scanned_pages.append(path.name)
        return "matched page" if path.name == "page_012.png" else "None"

    def evidence_judge(_question, _evidence):
        raise AssertionError("full_scan should not stop on sufficiency checks")

    workflow = build_docx_read_workflow(
        workspace,
        page_scanner=page_scanner,
        evidence_judge=evidence_judge,
    )

    result = workflow.invoke(
        {"file_ref": "file_001", "question": "summary?", "full_scan": True}
    )

    assert result["result"] == "page_12: matched page"
    assert result["scanned_pages"] == 25
    assert result["is_sufficient"] is True
    assert "page_025.png" in scanned_pages


def test_workflow_continues_when_page_evidence_is_not_sufficient(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(
        workspace,
        filename="report.docx",
        file_type="docx",
        page_count=25,
    )
    scanned_pages = []

    def page_scanner(path, _question):
        scanned_pages.append(path.name)
        if path.name in {"page_012.png", "page_023.png"}:
            return f"matched {path.name}"
        return "None"

    workflow = build_docx_read_workflow(
        workspace,
        page_scanner=page_scanner,
        evidence_judge=lambda _question, evidence: len(evidence) >= 2,
    )

    result = workflow.invoke({"file_ref": "file_001", "question": "summary?"})

    assert result["result"] == (
        "page_12: matched page_012.png\npage_23: matched page_023.png"
    )
    assert result["scanned_pages"] == 25


def test_read_tool_schema_exposes_full_scan() -> None:
    assert read_documents_module.read_docx_file.args["full_scan"] == {
        "default": 0,
        "title": "Full Scan",
        "type": "integer",
    }


def test_read_tool_passes_full_scan_to_workflow(monkeypatch) -> None:
    captured = {}

    class FakeWorkflow:
        def invoke(self, state):
            captured["state"] = state
            return {"result": "ok"}

    def fake_build_docx_read_workflow(workspace):
        captured["workspace"] = workspace
        return FakeWorkflow()

    monkeypatch.setattr(
        read_documents_module,
        "workspace_from_tool_runtime",
        lambda _runtime: "workspace",
    )
    monkeypatch.setattr(
        read_documents_module,
        "build_docx_read_workflow",
        fake_build_docx_read_workflow,
    )

    result = read_documents_module.read_docx_file.func(
        file_path="/report.docx",
        question="summary?",
        runtime=object(),
        full_scan=1,
    )

    assert result == "ok"
    assert captured["workspace"] == "workspace"
    assert captured["state"] == {
        "file_ref": "/report.docx",
        "question": "summary?",
        "full_scan": True,
    }


def test_read_tool_rejects_invalid_full_scan() -> None:
    result = read_documents_module.read_docx_file.func(
        file_path="/report.docx",
        question="summary?",
        runtime=object(),
        full_scan=2,
    )

    assert result == "full_scan must be 0 or 1."


def test_scan_page_uses_standard_image_block_with_filename(
    tmp_path,
    monkeypatch,
) -> None:
    captured = {}

    class FakeModel:
        def invoke(self, payload):
            captured["payload"] = payload
            return "None"

    def fake_llm_client(*, disable_streaming=False):
        assert disable_streaming is True
        return FakeModel()

    page_path = tmp_path / "page_001.png"
    page_path.write_bytes(b"png-bytes")
    monkeypatch.setattr(scan_module, "llm_client", fake_llm_client)

    result = scan_module.scan_page(
        page_path=page_path,
        question="question?",
        prompt="Question: {question}",
    )

    content = captured["payload"][0]["content"]
    image_block = content[1]
    assert result == "None"
    assert image_block == {
        "type": "image",
        "base64": base64.b64encode(b"png-bytes").decode("ascii"),
        "mime_type": "image/png",
        "filename": "page_001.png",
    }
    assert "image_url" not in image_block


def test_workflow_accepts_non_none_page_scan_as_evidence(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    _register_converted_file(workspace, filename="report.docx", file_type="docx")

    graph = build_docx_read_workflow(
        workspace,
        page_scanner=lambda _path, _question: "invalid",
        evidence_judge=lambda _question, _evidence: True,
    )

    result = graph.invoke({"file_ref": "file_001", "question": "summary?"})

    assert result["result"] == "page_1: invalid"


def test_workflow_edits_xlsx_and_registers_new_file(
    tmp_path,
    monkeypatch,
) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    source_path = workspace.files_dir / "book.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["A1"] = "old"
    workbook.save(source_path)
    workbook.close()
    converted_dir = workspace.converted_dir / "file_001"
    converted_dir.mkdir()
    (converted_dir / "manifest.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "book.xlsx",
                "source_path": str(source_path),
                "file_type": "xlsx",
                "pdf_path": str(converted_dir / "source.pdf"),
                "pages": [],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name="book.xlsx",
        file_type="xlsx",
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")

    def fake_build_upload_artifacts(**kwargs):
        target_dir = kwargs["converted_dir"]
        target_dir.mkdir(parents=True, exist_ok=True)
        (target_dir / "manifest.json").write_text(
            json.dumps(
                {
                    "file_id": kwargs["file_id"],
                    "source_filename": kwargs["source_path"].name,
                    "file_type": kwargs["file_type"],
                    "pages": [],
                    "status": "converted",
                }
            ),
            encoding="utf-8",
        )

    monkeypatch.setattr(file_registry, "build_upload_artifacts", fake_build_upload_artifacts)
    artifact = file_registry.resolve_artifact(
        workspace=workspace,
        file_ref="file_001",
        expected_file_type="xlsx",
    )
    session = XlsxSessionStore(workspace).create(artifact=artifact, instruction="edit")
    session.write_values(sheet="Summary", start_cell="A1", values=[["new"]])
    payload = commit_workbook(
        workspace=workspace,
        source_artifact=artifact,
        workbook_path=session.working_path,
        output_path="/book_edited.xlsx",
        summary="Updated Summary A1.",
        changed_items=session.changed_items(),
    )
    payload_json = json.dumps(payload, ensure_ascii=False)

    assert payload["file"]["file_id"] == "file_002"
    assert payload["file"]["download_url"].startswith("/api/fs/outputs/")
    assert ".outputs" not in payload_json
    assert ".converted" not in payload_json
    assert (workspace.files_dir / "book_edited.xlsx").is_file()

    edited = load_workbook(workspace.files_dir / "book_edited.xlsx")
    try:
        assert edited["Summary"]["A1"].value == "new"
    finally:
        edited.close()


def test_xlsx_read_workflow_uses_workbook_sheet_summary(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    source_path = workspace.files_dir / "book.xlsx"
    source_path.write_text("xlsx", encoding="utf-8")
    converted_dir = workspace.converted_dir / "file_001"
    summary_sheet_dir = converted_dir / "xlsx" / "sheets" / "sheet_0001"
    detail_sheet_dir = converted_dir / "xlsx" / "sheets" / "sheet_0002"
    summary_sheet_dir.mkdir(parents=True)
    detail_sheet_dir.mkdir(parents=True)
    (converted_dir / "manifest.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "book.xlsx",
                "source_path": str(source_path),
                "file_type": "xlsx",
                "pdf_path": str(converted_dir / "source.pdf"),
                "pages": [],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    (converted_dir / "workbook_index.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "book.xlsx",
                "sheet_count": 2,
                "sheets": [
                    {
                        "sheet_id": "sheet_0001",
                        "sheet_name": "Summary",
                        "index": 0,
                        "visible": True,
                        "used_range": "A1:B2",
                        "has_formulas": False,
                        "formula_count": 0,
                        "sheet_summary_path": str(summary_sheet_dir / "sheet.json"),
                    },
                    {
                        "sheet_id": "sheet_0002",
                        "sheet_name": "Detail",
                        "index": 1,
                        "visible": True,
                        "used_range": "A1:B2",
                        "has_formulas": False,
                        "formula_count": 0,
                        "sheet_summary_path": str(detail_sheet_dir / "sheet.json"),
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    (summary_sheet_dir / "sheet.json").write_text(
        json.dumps(
            {
                "sheet_id": "sheet_0001",
                "sheet_name": "Summary",
                "index": 0,
                "visible": True,
                "used_range": "A1:B2",
                "headers": ["Metric", "Value"],
                "sample_rows": [["Revenue", 10]],
                "formula_summary": "",
                "chart_summary": "",
                "text_summary": "Revenue summary",
                "pages": [
                    {
                        "page_number": 1,
                        "image_filename": "summary_page_001.png",
                        "image_path": str(summary_sheet_dir / "pages" / "page_001.png"),
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    (detail_sheet_dir / "sheet.json").write_text(
        json.dumps(
            {
                "sheet_id": "sheet_0002",
                "sheet_name": "Detail",
                "index": 1,
                "visible": True,
                "used_range": "A1:B2",
                "headers": ["Metric", "Value"],
                "sample_rows": [["Expense", 5]],
                "formula_summary": "",
                "chart_summary": "",
                "text_summary": "Expense summary",
                "pages": [
                    {
                        "page_number": 1,
                        "image_filename": "detail_page_001.png",
                        "image_path": str(detail_sheet_dir / "pages" / "page_001.png"),
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name="book.xlsx",
        file_type="xlsx",
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Revenue"
    summary["B2"] = 10
    detail = workbook.create_sheet("Detail")
    detail["A1"] = "Metric"
    detail["B1"] = "Value"
    detail["A2"] = "Expense"
    detail["B2"] = 5
    workbook.save(source_path)
    workbook.close()

    graph = build_xlsx_read_workflow(workspace)
    result = graph.invoke(
        {
            "file_ref": "file_001",
            "question": "Summary B1:B2 revenue?",
        }
    )
    payload = json.loads(result["result"])

    assert payload["workbook"]["sheet_count"] == 2
    assert payload["selected_range"]["range"]["sheet_name"] == "Summary"
    assert payload["selected_range"]["range"]["headers"] == ["Value"]
    assert payload["selected_range"]["range"]["rows"] == [{"Value": 10}]
    assert ".converted" not in result["result"]


def test_xlsx_read_workflow_rejects_missing_sheet_for_explicit_range(tmp_path) -> None:
    workspace = ensure_upload_workspace(tmp_path)
    source_path = workspace.files_dir / "book.xlsx"
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("Detail")
    workbook.save(source_path)
    workbook.close()
    converted_dir = workspace.converted_dir / "file_001"
    sheet_dir = converted_dir / "xlsx" / "sheets" / "sheet_0001"
    sheet_dir.mkdir(parents=True)
    (converted_dir / "manifest.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "book.xlsx",
                "source_path": str(source_path),
                "file_type": "xlsx",
                "pdf_path": str(converted_dir / "source.pdf"),
                "pages": [],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    (converted_dir / "workbook_index.json").write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": "book.xlsx",
                "sheet_count": 1,
                "sheets": [
                    {
                        "sheet_id": "sheet_0001",
                        "sheet_name": "Summary",
                        "index": 0,
                        "visible": True,
                        "used_range": "A1:B2",
                        "has_formulas": False,
                        "formula_count": 0,
                        "sheet_summary_path": str(sheet_dir / "sheet.json"),
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    (sheet_dir / "sheet.json").write_text(
        json.dumps(
            {
                "sheet_id": "sheet_0001",
                "sheet_name": "Summary",
                "index": 0,
                "visible": True,
                "used_range": "A1:B2",
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name="book.xlsx",
        file_type="xlsx",
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")

    graph = build_xlsx_read_workflow(workspace)

    with pytest.raises(ValueError, match="sheet name is required"):
        graph.invoke(
            {
                "file_ref": "file_001",
                "question": "B1:B2 revenue?",
            }
        )


def test_xlsx_worker_exposes_session_tools() -> None:
    tool_names = [tool.name for tool in build_xlsx_subagent()["tools"]]

    assert tool_names == [
        "start_xlsx_session",
        "inspect_xlsx_session",
        "load_xlsx_range",
        "profile_xlsx_dataframe",
        "preview_xlsx_dataframe",
        "transform_xlsx_dataframe",
        "write_xlsx_dataframe",
        "write_xlsx_values",
        "add_xlsx_formula",
        "export_xlsx_range",
        "export_xlsx_dataframe",
        "export_xlsx_detected_table_csv",
        "export_xlsx_dataframe_csv",
        "commit_xlsx_session",
        "discard_xlsx_session",
    ]


def _register_converted_file(
    workspace,
    *,
    filename: str,
    file_type: str,
    page_count: int = 1,
) -> None:
    source_path = workspace.files_dir / filename
    source_path.write_text("doc", encoding="utf-8")
    converted_dir = workspace.converted_dir / "file_001"
    converted_dir.mkdir()
    manifest_path = converted_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "file_id": "file_001",
                "source_filename": filename,
                "source_path": str(source_path),
                "file_type": file_type,
                "pdf_path": str(converted_dir / "source.pdf"),
                "pages": [
                    {
                        "page_number": index,
                        "image_filename": f"page_{index:03d}.png",
                        "image_path": str(converted_dir / "pages" / f"page_{index:03d}.png"),
                    }
                    for index in range(1, page_count + 1)
                ],
                "status": "converted",
            }
        ),
        encoding="utf-8",
    )
    registry = UploadRegistry(workspace.registry_path)
    registry.add_uploaded(
        file_id="file_001",
        visible_path=source_path,
        visible_name=filename,
        file_type=file_type,
        converted_dir=converted_dir,
    )
    registry.update_status("file_001", status="converted")
