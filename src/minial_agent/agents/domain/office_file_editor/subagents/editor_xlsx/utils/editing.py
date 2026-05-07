import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import range_boundaries


def apply_xlsx_edit(
    *,
    path: Path,
    operation: str,
    slots: dict[str, str],
    source_filename: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path)
    try:
        sheet_name = slots.get("SHEET") or workbook.sheetnames[0]
        if operation == "add_sheet":
            if sheet_name in workbook.sheetnames:
                raise ValueError(f"XLSX sheet already exists: {sheet_name}")
            worksheet = workbook.create_sheet(sheet_name)
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"XLSX sheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]

        cell = slots.get("CELL") or slots.get("START_CELL") or "A1"
        target_range = slots.get("RANGE") or _range_from_slots(slots, default_cell=cell)
        if operation in {"write_values", "write_formulas"}:
            changed_count = _write_xlsx_values_or_formulas(
                worksheet=worksheet,
                operation=operation,
                slots=slots,
                target_range=target_range,
            )
            new_value = slots.get("FORMULA_PATTERN") or slots.get("FORMULA")
            new_value = new_value or slots.get("VALUES") or slots.get("VALUE") or slots.get("TEXT")
        elif operation == "add_sheet":
            changed_count = 1
            new_value = slots.get("VALUE") or slots.get("TEXT")
            if new_value is not None:
                worksheet[cell] = new_value
        elif operation == "format_range":
            fill_color = slots.get("FILL")
            bold = slots.get("BOLD", "").lower() in {"1", "true", "yes"}
            changed_count = 0
            for target_cell in _iter_cells(worksheet, target_range):
                if fill_color:
                    target_cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=fill_color,
                    )
                    changed_count += 1
                if bold:
                    target_cell.font = Font(bold=True)
                    changed_count += 1
            new_value = json.dumps(
                {"range": target_range, "fill": fill_color, "bold": bold},
                ensure_ascii=False,
            )
        else:
            raise ValueError(f"Unsupported XLSX edit operation: {operation}")
        if changed_count == 0:
            raise ValueError("XLSX edit did not change the workbook.")
        workbook.save(path)
    finally:
        workbook.close()

    return [
        {
            "source_file": source_filename,
            "operation": operation,
            "sheet": sheet_name,
            "cell": cell,
            "range": target_range,
            "new_value": new_value,
            "changed_count": changed_count,
        }
    ]


def _write_xlsx_values_or_formulas(
    *,
    worksheet,
    operation: str,
    slots: dict[str, str],
    target_range: str,
) -> int:
    cells = list(_iter_cells(worksheet, target_range))
    if not cells:
        return 0

    if operation == "write_formulas":
        formula_pattern = slots.get("FORMULA_PATTERN")
        formula = slots.get("FORMULA")
        if formula_pattern:
            for target_cell in cells:
                value = formula_pattern.format(
                    row=target_cell.row,
                    prev_row=target_cell.row - 1,
                )
                target_cell.value = _formula_value(value)
            return len(cells)
        if formula is None:
            raise ValueError("XLSX write_formulas requires FORMULA or FORMULA_PATTERN.")
        formula = _formula_value(formula)
        for target_cell in cells:
            target_cell.value = formula
        return len(cells)

    if "VALUES" in slots:
        values = _parse_values(slots["VALUES"])
        min_col, min_row, max_col, max_row = range_boundaries(target_range)
        row_count = max_row - min_row + 1
        column_count = max_col - min_col + 1
        if len(values) != row_count or any(len(row) != column_count for row in values):
            raise ValueError("XLSX VALUES shape does not match target range.")
        changed_count = 0
        for row_offset, row in enumerate(values):
            for column_offset, value in enumerate(row):
                worksheet.cell(
                    row=min_row + row_offset,
                    column=min_col + column_offset,
                    value=value,
                )
                changed_count += 1
        return changed_count

    value = slots.get("VALUE") or slots.get("TEXT")
    if value is None:
        raise ValueError("XLSX write_values requires VALUE, TEXT, or VALUES.")
    for target_cell in cells:
        target_cell.value = value
    return len(cells)


def _range_from_slots(slots: dict[str, str], *, default_cell: str) -> str:
    if slots.get("START_CELL") and slots.get("END_CELL"):
        return f"{slots['START_CELL']}:{slots['END_CELL']}"
    return slots.get("RANGE") or default_cell


def _iter_cells(worksheet, target_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(target_range)
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            yield cell


def _parse_values(raw_values: str) -> list[list[str]]:
    return [
        [value.strip() for value in row.split(",")]
        for row in raw_values.split("|")
    ]


def _formula_value(value: str) -> str:
    return value if value.startswith("=") else f"={value}"
