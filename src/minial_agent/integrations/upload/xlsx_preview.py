from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def build_xlsx_preview(source_path: Path) -> dict[str, Any]:
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    value_workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            value_sheet = value_workbook[worksheet.title]
            max_row = max(worksheet.max_row or 1, 1)
            max_column = max(worksheet.max_column or 1, 1)
            cells = []

            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    display_value = value_sheet[cell.coordinate].value
                    if value is None and display_value is None:
                        continue
                    cells.append(
                        {
                            "address": cell.coordinate,
                            "row": cell.row,
                            "column": cell.column,
                            "value": json_value(display_value if display_value is not None else value),
                            "formula": value if isinstance(value, str) and value.startswith("=") else None,
                            "style": cell_style(cell),
                        }
                    )

            sheets.append(
                {
                    "id": worksheet.title,
                    "name": worksheet.title,
                    "visible": worksheet.sheet_state == "visible",
                    "index": len(sheets),
                    "used_range": worksheet.calculate_dimension(),
                    "row_count": max_row,
                    "column_count": max_column,
                    "columns": [
                        {
                            "index": index,
                            "label": get_column_letter(index),
                            "width": column_width(worksheet.column_dimensions[get_column_letter(index)].width),
                        }
                        for index in range(1, max_column + 1)
                    ],
                    "rows": [
                        {
                            "index": index,
                            "height": row_height(worksheet.row_dimensions[index].height),
                        }
                        for index in range(1, max_row + 1)
                    ],
                    "merged_ranges": [str(range_) for range_ in worksheet.merged_cells.ranges],
                    "cells": cells,
                }
            )

        return {
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()
        value_workbook.close()


def json_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def cell_style(cell) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    style: dict[str, Any] = {
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
    }
    if font.color and font.color.type == "rgb" and font.color.rgb:
        style["color"] = argb_to_css(font.color.rgb)
    if fill.fill_type == "solid" and fill.fgColor.type == "rgb" and fill.fgColor.rgb:
        style["background"] = argb_to_css(fill.fgColor.rgb)
    return style


def argb_to_css(value: str) -> str:
    normalized = value[-6:]
    return f"#{normalized}"


def column_width(width: float | None) -> int:
    return max(48, min(int((width or 8.43) * 8), 240))


def row_height(height: float | None) -> int:
    return max(24, min(int((height or 15) * 1.4), 120))
