import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from minial_agent.integrations.upload.conversion import ConversionError
from minial_agent.integrations.xlsx.workbook import inspect_workbook as inspect_xlsx_workbook


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def inspect_workbook(source_path: Path) -> list[dict[str, object]]:
    try:
        return [
            _upload_sheet_entry(sheet)
            for sheet in inspect_xlsx_workbook(source_path, filename=source_path.name).sheets
        ]
    except Exception:
        return _inspect_workbook_with_zip(source_path)


def _upload_sheet_entry(sheet) -> dict[str, object]:
    payload = sheet.to_dict()
    formulas = payload.get("formulas", [])
    return {
        "sheet_name": payload["sheet_name"],
        "visible": payload["visible"],
        "used_range": payload["used_range"],
        "headers": payload["headers"],
        "sample_rows": payload["sample_rows"],
        "has_formulas": payload["has_formulas"],
        "formula_count": payload["formula_count"],
        "has_tables": payload["has_tables"],
        "has_charts": payload["has_charts"],
        "formula_summary": formula_summary(formulas if isinstance(formulas, list) else []),
        "chart_summary": (
            f"{payload['chart_count']} chart(s) exist." if payload["chart_count"] else "No charts."
        ),
        "text_summary": text_summary(
            str(payload["sheet_name"]),
            str(payload["used_range"]),
            [str(header) for header in payload["headers"]],
        ),
    }


def _inspect_workbook_with_openpyxl(source_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            formulas = [
                cell.coordinate
                for row in worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(6, worksheet.max_row),
                    values_only=True,
                )
            )
            headers = [str(value) for value in rows[0] if value is not None] if rows else []
            sample_rows = [
                [value for value in row]
                for row in rows[1:6]
                if any(value is not None for value in row)
            ]
            table_count = len(getattr(worksheet, "tables", {}))
            chart_count = len(getattr(worksheet, "_charts", []))
            used_range = worksheet.calculate_dimension()
            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "visible": worksheet.sheet_state == "visible",
                    "used_range": used_range,
                    "headers": headers,
                    "sample_rows": sample_rows,
                    "has_formulas": bool(formulas),
                    "formula_count": len(formulas),
                    "has_tables": table_count > 0,
                    "has_charts": chart_count > 0,
                    "formula_summary": formula_summary(formulas),
                    "chart_summary": (
                        f"{chart_count} chart(s) exist." if chart_count else "No charts."
                    ),
                    "text_summary": text_summary(worksheet.title, used_range, headers),
                }
            )
        return sheets
    finally:
        workbook.close()


def _inspect_workbook_with_zip(source_path: Path) -> list[dict[str, object]]:
    with zipfile.ZipFile(source_path) as workbook:
        workbook_xml = ElementTree.fromstring(workbook.read("xl/workbook.xml"))
        rels = read_workbook_relationships(workbook)
        sheets = []

        for sheet in workbook_xml.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
            relationship_id = sheet.attrib.get(f"{{{REL_NS}}}id")
            target = rels.get(relationship_id or "")
            worksheet_xml = ElementTree.fromstring(workbook.read(worksheet_path(target)))
            formulas = worksheet_xml.findall(f".//{{{SPREADSHEET_NS}}}f")
            dimension = worksheet_xml.find(f".//{{{SPREADSHEET_NS}}}dimension")
            state = sheet.attrib.get("state", "visible")
            sheets.append(
                {
                    "sheet_name": sheet.attrib.get("name", ""),
                    "visible": state == "visible",
                    "used_range": dimension.attrib.get("ref", "") if dimension is not None else "",
                    "headers": [],
                    "sample_rows": [],
                    "has_formulas": bool(formulas),
                    "formula_count": len(formulas),
                    "has_tables": False,
                    "has_charts": False,
                    "formula_summary": formula_summary(
                        [formula.text or "" for formula in formulas]
                    ),
                    "chart_summary": "No charts.",
                    "text_summary": "",
                }
            )

        return sheets


def formula_summary(formulas: list[str]) -> str:
    if not formulas:
        return "No formulas."
    examples = ", ".join(formulas[:5])
    return f"{len(formulas)} formula(s). Examples: {examples}"


def text_summary(sheet_name: str, used_range: str, headers: list[str]) -> str:
    header_text = ", ".join(headers[:8])
    if header_text:
        return f"{sheet_name} sheet uses {used_range} with headers: {header_text}."
    return f"{sheet_name} sheet uses {used_range}."


def read_workbook_relationships(workbook: zipfile.ZipFile) -> dict[str, str]:
    rels_xml = ElementTree.fromstring(workbook.read("xl/_rels/workbook.xml.rels"))
    rels = {}
    for relationship in rels_xml.findall(f".//{{{PACKAGE_REL_NS}}}Relationship"):
        rels[relationship.attrib.get("Id", "")] = relationship.attrib.get("Target", "")
    return rels


def worksheet_path(target: str | None) -> str:
    if not target:
        raise ConversionError("Workbook sheet relationship is missing a target")
    normalized = re.sub(r"^/+", "", target)
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"
