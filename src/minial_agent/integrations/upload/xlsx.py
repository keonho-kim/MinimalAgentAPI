import json
import re
import tempfile
import zipfile
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from minial_agent.integrations.upload.conversion import ConversionError, convert_to_pdf, render_pdf_pages
from minial_agent.integrations.xlsx.workbook import inspect_workbook as inspect_xlsx_workbook


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def build_xlsx_artifacts(
    *,
    source_path: Path,
    file_id: str,
    converted_dir: Path,
    cache_dir: Path,
) -> None:
    sheets = inspect_workbook(source_path)
    xlsx_dir = converted_dir / "xlsx"
    sheets_dir = xlsx_dir / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)

    workbook_entries = []
    for index, sheet in enumerate(sheets):
        sheet_id = f"sheet_{index + 1:04d}"
        sheet_dir = sheets_dir / sheet_id
        sheet_dir.mkdir(parents=True, exist_ok=True)
        sheet_summary_path = sheet_dir / "sheet.json"
        sheet_entry = {
            "sheet_id": sheet_id,
            "sheet_name": sheet["sheet_name"],
            "index": index,
            "visible": sheet["visible"],
            "used_range": sheet["used_range"],
            "has_formulas": sheet["has_formulas"],
            "formula_count": sheet["formula_count"],
            "has_tables": sheet.get("has_tables", False),
            "has_charts": sheet.get("has_charts", False),
            "sheet_summary_path": str(sheet_summary_path),
        }

        sheet_summary = dict(sheet_entry)
        sheet_summary["headers"] = sheet.get("headers", [])
        sheet_summary["sample_rows"] = sheet.get("sample_rows", [])
        sheet_summary["formula_summary"] = sheet.get("formula_summary", "")
        sheet_summary["chart_summary"] = sheet.get("chart_summary", "")
        sheet_summary["text_summary"] = sheet.get("text_summary", "")
        try:
            pages = _render_single_sheet(
                source_path=source_path,
                sheet_index=index,
                sheet_dir=sheet_dir,
                cache_dir=cache_dir,
            )
            sheet_summary["pages"] = [
                {
                    "page_number": page.page_number,
                    "image_filename": page.image_filename,
                    "image_path": page.image_path,
                }
                for page in pages
            ]
        except Exception as exc:
            sheet_summary["conversion_error"] = str(exc)

        sheet_summary_path.write_text(
            json.dumps(sheet_summary, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        workbook_entries.append(sheet_entry)

    workbook_index = {
        "file_id": file_id,
        "source_filename": source_path.name,
        "sheet_count": len(workbook_entries),
        "sheets": workbook_entries,
    }
    (converted_dir / "workbook_index.json").write_text(
        json.dumps(workbook_index, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def inspect_workbook(source_path: Path) -> list[dict[str, object]]:
    try:
        return [
            _upload_sheet_entry(sheet)
            for sheet in inspect_xlsx_workbook(source_path, filename=source_path.name).sheets
        ]
    except Exception:
        return _inspect_workbook_with_zip(source_path)


def _upload_sheet_entry(sheet) -> dict[str, object]:
    payload = sheet.to_dict()
    formulas = payload.get("formulas", [])
    return {
        "sheet_name": payload["sheet_name"],
        "visible": payload["visible"],
        "used_range": payload["used_range"],
        "headers": payload["headers"],
        "sample_rows": payload["sample_rows"],
        "has_formulas": payload["has_formulas"],
        "formula_count": payload["formula_count"],
        "has_tables": payload["has_tables"],
        "has_charts": payload["has_charts"],
        "formula_summary": _formula_summary(formulas if isinstance(formulas, list) else []),
        "chart_summary": (
            f"{payload['chart_count']} chart(s) exist." if payload["chart_count"] else "No charts."
        ),
        "text_summary": _text_summary(
            str(payload["sheet_name"]),
            str(payload["used_range"]),
            [str(header) for header in payload["headers"]],
        ),
    }


def build_xlsx_preview(source_path: Path) -> dict[str, Any]:
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    value_workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            value_sheet = value_workbook[worksheet.title]
            max_row = max(worksheet.max_row or 1, 1)
            max_column = max(worksheet.max_column or 1, 1)
            cells = []

            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    display_value = value_sheet[cell.coordinate].value
                    if value is None and display_value is None:
                        continue
                    cells.append(
                        {
                            "address": cell.coordinate,
                            "row": cell.row,
                            "column": cell.column,
                            "value": _json_value(display_value if display_value is not None else value),
                            "formula": value if isinstance(value, str) and value.startswith("=") else None,
                            "style": _cell_style(cell),
                        }
                    )

            sheets.append(
                {
                    "id": worksheet.title,
                    "name": worksheet.title,
                    "visible": worksheet.sheet_state == "visible",
                    "index": len(sheets),
                    "used_range": worksheet.calculate_dimension(),
                    "row_count": max_row,
                    "column_count": max_column,
                    "columns": [
                        {
                            "index": index,
                            "label": get_column_letter(index),
                            "width": _column_width(worksheet.column_dimensions[get_column_letter(index)].width),
                        }
                        for index in range(1, max_column + 1)
                    ],
                    "rows": [
                        {
                            "index": index,
                            "height": _row_height(worksheet.row_dimensions[index].height),
                        }
                        for index in range(1, max_row + 1)
                    ],
                    "merged_ranges": [str(range_) for range_ in worksheet.merged_cells.ranges],
                    "cells": cells,
                }
            )

        return {
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()
        value_workbook.close()


def _inspect_workbook_with_openpyxl(source_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            formulas = [
                cell.coordinate
                for row in worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ]
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(6, worksheet.max_row),
                    values_only=True,
                )
            )
            headers = [str(value) for value in rows[0] if value is not None] if rows else []
            sample_rows = [
                [value for value in row]
                for row in rows[1:6]
                if any(value is not None for value in row)
            ]
            table_count = len(getattr(worksheet, "tables", {}))
            chart_count = len(getattr(worksheet, "_charts", []))
            used_range = worksheet.calculate_dimension()
            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "visible": worksheet.sheet_state == "visible",
                    "used_range": used_range,
                    "headers": headers,
                    "sample_rows": sample_rows,
                    "has_formulas": bool(formulas),
                    "formula_count": len(formulas),
                    "has_tables": table_count > 0,
                    "has_charts": chart_count > 0,
                    "formula_summary": _formula_summary(formulas),
                    "chart_summary": (
                        f"{chart_count} chart(s) exist." if chart_count else "No charts."
                    ),
                    "text_summary": _text_summary(worksheet.title, used_range, headers),
                }
            )
        return sheets
    finally:
        workbook.close()


def _inspect_workbook_with_zip(source_path: Path) -> list[dict[str, object]]:
    with zipfile.ZipFile(source_path) as workbook:
        workbook_xml = ElementTree.fromstring(workbook.read("xl/workbook.xml"))
        rels = _read_workbook_relationships(workbook)
        sheets = []

        for sheet in workbook_xml.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
            relationship_id = sheet.attrib.get(f"{{{REL_NS}}}id")
            target = rels.get(relationship_id or "")
            worksheet_path = _worksheet_path(target)
            worksheet_xml = ElementTree.fromstring(workbook.read(worksheet_path))
            formulas = worksheet_xml.findall(f".//{{{SPREADSHEET_NS}}}f")
            dimension = worksheet_xml.find(f".//{{{SPREADSHEET_NS}}}dimension")
            state = sheet.attrib.get("state", "visible")
            sheets.append(
                {
                    "sheet_name": sheet.attrib.get("name", ""),
                    "visible": state == "visible",
                    "used_range": dimension.attrib.get("ref", "") if dimension is not None else "",
                    "headers": [],
                    "sample_rows": [],
                    "has_formulas": bool(formulas),
                    "formula_count": len(formulas),
                    "has_tables": False,
                    "has_charts": False,
                    "formula_summary": _formula_summary(
                        [formula.text or "" for formula in formulas]
                    ),
                    "chart_summary": "No charts.",
                    "text_summary": "",
                }
            )

        return sheets


def _formula_summary(formulas: list[str]) -> str:
    if not formulas:
        return "No formulas."
    examples = ", ".join(formulas[:5])
    return f"{len(formulas)} formula(s). Examples: {examples}"


def _text_summary(sheet_name: str, used_range: str, headers: list[str]) -> str:
    header_text = ", ".join(headers[:8])
    if header_text:
        return f"{sheet_name} sheet uses {used_range} with headers: {header_text}."
    return f"{sheet_name} sheet uses {used_range}."


def _json_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def _cell_style(cell) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    style: dict[str, Any] = {
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
    }
    if font.color and font.color.type == "rgb" and font.color.rgb:
        style["color"] = _argb_to_css(font.color.rgb)
    if fill.fill_type == "solid" and fill.fgColor.type == "rgb" and fill.fgColor.rgb:
        style["background"] = _argb_to_css(fill.fgColor.rgb)
    return style


def _argb_to_css(value: str) -> str:
    normalized = value[-6:]
    return f"#{normalized}"


def _column_width(width: float | None) -> int:
    return max(48, min(int((width or 8.43) * 8), 240))


def _row_height(height: float | None) -> int:
    return max(24, min(int((height or 15) * 1.4), 120))


def _render_single_sheet(
    *,
    source_path: Path,
    sheet_index: int,
    sheet_dir: Path,
    cache_dir: Path,
):
    pages_dir = sheet_dir / "pages"
    with tempfile.TemporaryDirectory(dir=cache_dir) as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        single_sheet_path = temp_dir / source_path.name
        _copy_with_only_sheet_visible(source_path, single_sheet_path, sheet_index)
        pdf_dir = temp_dir / "pdf"
        pdf_path = temp_dir / "source.pdf"
        convert_to_pdf(single_sheet_path, pdf_dir, pdf_path)
        return render_pdf_pages(pdf_path, pages_dir)


def _copy_with_only_sheet_visible(
    source_path: Path,
    target_path: Path,
    visible_sheet_index: int,
) -> None:
    with zipfile.ZipFile(source_path) as source:
        with zipfile.ZipFile(target_path, "w", zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    data = _set_visible_sheet(data, visible_sheet_index)
                target.writestr(item, data)


def _set_visible_sheet(workbook_xml: bytes, visible_sheet_index: int) -> bytes:
    root = ElementTree.fromstring(workbook_xml)
    sheets = root.findall(f".//{{{SPREADSHEET_NS}}}sheet")
    for index, sheet in enumerate(sheets):
        if index == visible_sheet_index:
            sheet.attrib.pop("state", None)
        else:
            sheet.attrib["state"] = "hidden"
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _read_workbook_relationships(workbook: zipfile.ZipFile) -> dict[str, str]:
    rels_xml = ElementTree.fromstring(workbook.read("xl/_rels/workbook.xml.rels"))
    rels = {}
    for relationship in rels_xml.findall(f".//{{{PACKAGE_REL_NS}}}Relationship"):
        rels[relationship.attrib.get("Id", "")] = relationship.attrib.get("Target", "")
    return rels


def _worksheet_path(target: str | None) -> str:
    if not target:
        raise ConversionError("Workbook sheet relationship is missing a target")
    normalized = re.sub(r"^/+", "", target)
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"
