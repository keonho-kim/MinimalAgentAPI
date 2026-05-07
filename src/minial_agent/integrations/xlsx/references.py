from pathlib import Path
from typing import Iterable, Protocol

from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.exceptions import CellCoordinatesException

from minial_agent.integrations.xlsx.errors import XlsxRangeError


class XlsxWorkspace(Protocol):
    files_dir: Path


def resolve_sheet_name(sheet_names: Iterable[str], requested_sheet: str) -> str:
    requested = requested_sheet.strip().strip("'")
    if not requested:
        raise XlsxRangeError("Sheet name is required.")
    names = list(sheet_names)
    for name in names:
        if name == requested:
            return name
    for name in names:
        if name.lower() == requested.lower():
            return name
    raise XlsxRangeError(f"XLSX sheet not found: {requested_sheet}")


def normalize_a1_range(reference: str) -> str:
    ref = reference.strip()
    if not ref:
        raise XlsxRangeError("XLSX range is required.")
    try:
        if ":" not in ref:
            coordinate_from_string(ref)
            return f"{ref}:{ref}"
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except (CellCoordinatesException, ValueError) as exc:
        raise XlsxRangeError(f"Invalid XLSX range: {reference}") from exc
    if min_col > max_col or min_row > max_row:
        raise XlsxRangeError(f"Invalid XLSX range: {reference}")
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def parse_sheet_range(reference: str) -> tuple[str | None, str]:
    raw = reference.strip()
    if "!" not in raw:
        return None, normalize_a1_range(raw)
    sheet, range_ref = raw.rsplit("!", 1)
    return sheet.strip().strip("'"), normalize_a1_range(range_ref)


def start_cell_bounds(start_cell: str, row_count: int, column_count: int) -> tuple[int, int, int, int]:
    column, row = coordinate_from_string(start_cell.strip())
    min_col = column_index_from_string(column)
    min_row = int(row)
    return min_col, min_row, min_col + column_count - 1, min_row + row_count - 1


def target_range_from_start(start_cell: str, row_count: int, column_count: int) -> str:
    min_col, min_row, max_col, max_row = start_cell_bounds(start_cell, row_count, column_count)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def resolve_output_path(
    workspace: XlsxWorkspace,
    output_path: str,
    *,
    allowed_extensions: set[str],
) -> Path:
    from minial_agent.integrations.upload.visibility import public_virtual_to_physical

    path = public_virtual_to_physical(workspace.files_dir, output_path)
    suffix = path.suffix.lower()
    if suffix not in allowed_extensions:
        allowed = ", ".join(sorted(allowed_extensions))
        raise XlsxRangeError(f"Output path must end with one of: {allowed}")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def resolve_output_filename(
    workspace: XlsxWorkspace,
    output_filename: str,
    *,
    extension: str,
) -> Path:
    suffix = extension if extension.startswith(".") else f".{extension}"
    raw = output_filename.strip()
    if not raw:
        raise XlsxRangeError("Output filename is required.")
    if not Path(raw).suffix:
        raw = f"{raw}{suffix}"
    return resolve_output_path(workspace, raw, allowed_extensions={suffix.lower()})
