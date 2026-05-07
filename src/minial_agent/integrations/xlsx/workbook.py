from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from minial_agent.integrations.xlsx.errors import XlsxRangeError
from minial_agent.integrations.xlsx.models import SheetInspection, WorkbookInspection
from minial_agent.integrations.xlsx.references import resolve_sheet_name


def open_workbook(path: Path, *, data_only: bool = False) -> Workbook:
    return load_workbook(path, read_only=False, data_only=data_only)


def get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    resolved = resolve_sheet_name(workbook.sheetnames, sheet_name)
    return workbook[resolved]


def inspect_workbook(path: Path, *, filename: str | None = None) -> WorkbookInspection:
    workbook = open_workbook(path, data_only=False)
    value_workbook = open_workbook(path, data_only=True)
    try:
        sheets = [
            inspect_sheet(worksheet, value_workbook[worksheet.title], index)
            for index, worksheet in enumerate(workbook.worksheets)
        ]
        return WorkbookInspection(
            filename=filename or path.name,
            sheet_count=len(sheets),
            sheets=sheets,
        )
    finally:
        workbook.close()
        value_workbook.close()


def inspect_sheet(worksheet: Worksheet, value_sheet: Worksheet, index: int) -> SheetInspection:
    used_range = worksheet.calculate_dimension()
    headers, sample_rows = _headers_and_samples(value_sheet)
    formulas = [
        cell.coordinate
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    tables = _tables(worksheet)
    candidates = [item["range"] for item in tables]
    if used_range and used_range not in {"A1", "A1:A1"}:
        candidates.append(used_range)
    return SheetInspection(
        sheet_name=worksheet.title,
        index=index,
        visible=worksheet.sheet_state == "visible",
        used_range=used_range,
        row_count=max(worksheet.max_row or 1, 1),
        column_count=max(worksheet.max_column or 1, 1),
        headers=headers,
        sample_rows=sample_rows,
        formulas=formulas[:200],
        formula_count=len(formulas),
        tables=tables,
        chart_count=len(getattr(worksheet, "_charts", [])),
        merged_ranges=[str(item) for item in worksheet.merged_cells.ranges],
        candidate_ranges=candidates,
    )


def sheet_names(path: Path) -> list[str]:
    workbook = open_workbook(path, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _headers_and_samples(worksheet: Worksheet) -> tuple[list[str], list[list[Any]]]:
    rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(max(worksheet.max_row or 1, 1), 8),
            values_only=True,
        )
    )
    non_empty_rows = [row for row in rows if any(value is not None for value in row)]
    if not non_empty_rows:
        return [], []
    headers = [str(value) for value in non_empty_rows[0] if value is not None]
    samples = [
        [_json_value(value) for value in row]
        for row in non_empty_rows[1:6]
    ]
    return headers, samples


def _tables(worksheet: Worksheet) -> list[dict[str, Any]]:
    tables = []
    table_values = getattr(worksheet.tables, "values", None)
    if not callable(table_values):
        return tables
    for table in table_values():
        ref = str(getattr(table, "ref", ""))
        if not ref:
            continue
        tables.append(
            {
                "name": str(getattr(table, "name", "")),
                "display_name": str(getattr(table, "displayName", "")),
                "range": ref,
            }
        )
    return tables


def _json_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def require_sheet(path: Path, sheet_name: str) -> str:
    names = sheet_names(path)
    if not names:
        raise XlsxRangeError("Workbook has no sheets.")
    return resolve_sheet_name(names, sheet_name)
