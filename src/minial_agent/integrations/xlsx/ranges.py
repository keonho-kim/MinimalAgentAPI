from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, range_boundaries

from minial_agent.integrations.xlsx.errors import XlsxRangeError
from minial_agent.integrations.xlsx.models import RangeReadResult
from minial_agent.integrations.xlsx.references import normalize_a1_range, target_range_from_start
from minial_agent.integrations.xlsx.workbook import get_sheet, open_workbook


def read_range(path: Path, sheet: str, range_ref: str, *, header: bool = True) -> RangeReadResult:
    normalized = normalize_a1_range(range_ref)
    workbook = open_workbook(path, data_only=True)
    try:
        worksheet = get_sheet(workbook, sheet)
        values = [
            [_json_value(cell.value) for cell in row]
            for row in worksheet[normalized]
        ]
        if not values:
            raise XlsxRangeError(f"XLSX range is empty: {sheet}!{range_ref}")
        headers = _headers(values, header=header)
        data_values = values[1:] if header else values
        rows = [
            {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
            for row in data_values
        ]
        return RangeReadResult(
            sheet_name=worksheet.title,
            range=normalized,
            headers=headers,
            rows=rows,
            values=values,
        )
    finally:
        workbook.close()


def range_to_dataframe(path: Path, sheet: str, range_ref: str, *, header: bool = True) -> pd.DataFrame:
    result = read_range(path, sheet, range_ref, header=header)
    return pd.DataFrame(result.rows, columns=result.headers)


def write_values(path: Path, sheet: str, start_cell: str, values: list[list[Any]]) -> dict[str, Any]:
    if not values:
        raise XlsxRangeError("values must not be empty.")
    workbook = open_workbook(path, data_only=False)
    try:
        worksheet = get_sheet(workbook, sheet)
        target = target_range_from_start(start_cell, len(values), max(len(row) for row in values))
        min_col, min_row, _max_col, _max_row = range_boundaries(target)
        changed = 0
        for row_index, row in enumerate(values):
            for column_index, value in enumerate(row):
                worksheet.cell(row=min_row + row_index, column=min_col + column_index).value = value
                changed += 1
        workbook.save(path)
        return {"sheet": worksheet.title, "range": target, "changed_count": changed}
    finally:
        workbook.close()


def write_formula(
    path: Path,
    sheet: str,
    cell: str,
    formula: str,
    *,
    fill_range: str | None = None,
) -> dict[str, Any]:
    if not formula.startswith("="):
        formula = f"={formula}"
    workbook = open_workbook(path, data_only=False)
    try:
        worksheet = get_sheet(workbook, sheet)
        target = normalize_a1_range(fill_range or cell)
        min_col, min_row, max_col, max_row = range_boundaries(target)
        changed = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                address = f"{get_column_letter(col)}{row}"
                worksheet[address] = (
                    formula
                    .replace("{row}", str(row))
                    .replace("{column}", get_column_letter(col))
                    .replace("{cell}", address)
                )
                changed += 1
        workbook.save(path)
        return {"sheet": worksheet.title, "range": target, "changed_count": changed}
    finally:
        workbook.close()


def write_dataframe(
    path: Path,
    sheet: str,
    start_cell: str,
    dataframe: pd.DataFrame,
    *,
    include_header: bool = True,
) -> dict[str, Any]:
    values: list[list[Any]] = []
    if include_header:
        values.append([str(column) for column in dataframe.columns])
    values.extend(dataframe.where(pd.notna(dataframe), None).values.tolist())
    return write_values(path, sheet, start_cell, values)


def copy_range_to_workbook(source_path: Path, sheet: str, range_ref: str, output_path: Path) -> dict[str, Any]:
    result = read_range(source_path, sheet, range_ref, header=False)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = result.sheet_name[:31] or "Sheet1"
    for row_index, row in enumerate(result.values, start=1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value
    workbook.save(output_path)
    workbook.close()
    return {
        "output_path": output_path.name,
        "sheet": result.sheet_name,
        "range": result.range,
        "row_count": len(result.values),
        "column_count": len(result.values[0]) if result.values else 0,
    }


def _headers(values: list[list[Any]], *, header: bool) -> list[str]:
    width = max(len(row) for row in values)
    if not header:
        return [f"Column {index}" for index in range(1, width + 1)]
    return _dedupe_headers([
        str(value).strip() if value not in {None, ""} else f"Column {index}"
        for index, value in enumerate(values[0], start=1)
    ])


def _dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for header in headers:
        base = header or "Column"
        count = counts.get(base, 0) + 1
        counts[base] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def _json_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value
