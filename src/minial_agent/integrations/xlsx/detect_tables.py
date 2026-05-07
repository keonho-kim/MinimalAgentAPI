from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from minial_agent.integrations.xlsx.errors import XlsxRangeError
from minial_agent.integrations.xlsx.workbook import get_sheet, open_workbook


@dataclass(frozen=True)
class DetectedTable:
    sheet: str
    range: str
    header_row: int
    data_start_row: int
    row_count: int
    column_count: int
    headers: list[str]
    confidence: float
    evidence: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "range": self.range,
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "headers": self.headers,
            "confidence": self.confidence,
            "evidence": self.evidence,
        }


def detect_table(path, *, sheet: str | None = None) -> DetectedTable:
    workbook = open_workbook(path, data_only=True)
    try:
        worksheets = [get_sheet(workbook, sheet)] if sheet else workbook.worksheets
        candidates = [_detect_sheet_table(worksheet) for worksheet in worksheets]
        candidates = [candidate for candidate in candidates if candidate is not None]
        if not candidates:
            raise XlsxRangeError("Could not detect a tabular data region in the XLSX workbook.")
        return max(candidates, key=lambda candidate: (candidate.confidence, candidate.row_count, candidate.column_count))
    finally:
        workbook.close()


def _detect_sheet_table(worksheet: Worksheet) -> DetectedTable | None:
    min_row, max_row, min_col, max_col = _non_empty_bounds(worksheet)
    if min_row is None:
        return None

    best: DetectedTable | None = None
    for row_index in range(min_row, max_row + 1):
        header_cells = [worksheet.cell(row_index, column).value for column in range(min_col, max_col + 1)]
        header_cols = [
            min_col + index
            for index, value in enumerate(header_cells)
            if _is_header_value(value)
        ]
        if len(header_cols) < 2:
            continue
        first_col = min(header_cols)
        last_col = max(header_cols)
        headers = [
            str(worksheet.cell(row_index, column).value).strip()
            for column in range(first_col, last_col + 1)
        ]
        if any(not header for header in headers):
            continue

        data_start = row_index + 1
        data_end = _data_end_row(worksheet, data_start, max_row, first_col, last_col)
        if data_end < data_start:
            continue
        data_rows = data_end - data_start + 1
        column_count = last_col - first_col + 1
        density = _data_density(worksheet, data_start, data_end, first_col, last_col)
        confidence = _confidence(headers=headers, data_rows=data_rows, density=density)
        candidate = DetectedTable(
            sheet=worksheet.title,
            range=f"{get_column_letter(first_col)}{row_index}:{get_column_letter(last_col)}{data_end}",
            header_row=row_index,
            data_start_row=data_start,
            row_count=data_rows,
            column_count=column_count,
            headers=headers,
            confidence=confidence,
            evidence=(
                f"Detected {column_count} headers on row {row_index} with "
                f"{data_rows} dense data rows below."
            ),
        )
        if best is None or (candidate.confidence, candidate.row_count, candidate.column_count) > (
            best.confidence,
            best.row_count,
            best.column_count,
        ):
            best = candidate
    return best


def _non_empty_bounds(worksheet: Worksheet) -> tuple[int | None, int, int, int]:
    rows = []
    columns = []
    for row in worksheet.iter_rows():
        for cell in row:
            if _has_value(cell.value):
                rows.append(cell.row)
                columns.append(cell.column)
    if not rows:
        return None, 0, 0, 0
    return min(rows), max(rows), min(columns), max(columns)


def _data_end_row(
    worksheet: Worksheet,
    data_start: int,
    max_row: int,
    first_col: int,
    last_col: int,
) -> int:
    data_end = data_start - 1
    blank_streak = 0
    for row_index in range(data_start, max_row + 1):
        values = [
            worksheet.cell(row_index, column).value
            for column in range(first_col, last_col + 1)
        ]
        non_empty = sum(1 for value in values if _has_value(value))
        if non_empty == 0:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        if non_empty / (last_col - first_col + 1) >= 0.4:
            data_end = row_index
    return data_end


def _data_density(
    worksheet: Worksheet,
    data_start: int,
    data_end: int,
    first_col: int,
    last_col: int,
) -> float:
    total = (data_end - data_start + 1) * (last_col - first_col + 1)
    if total <= 0:
        return 0
    present = 0
    for row_index in range(data_start, data_end + 1):
        for column in range(first_col, last_col + 1):
            if _has_value(worksheet.cell(row_index, column).value):
                present += 1
    return present / total


def _confidence(*, headers: list[str], data_rows: int, density: float) -> float:
    header_score = min(len(headers) / 8, 1.0)
    row_score = min(data_rows / 20, 1.0)
    return round((header_score * 0.35) + (row_score * 0.35) + (density * 0.30), 3)


def _is_header_value(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    if len(text.split()) > 8:
        return False
    return True


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""
